"""
Exportador de Excel con formato ejecutivo para presentacion.
"""
from __future__ import annotations

from io import BytesIO
from typing import Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExportadorExcel:
    """Exporta resultados a Excel con formato ejecutivo."""

    ESTADO_COL = "Estado de Auditoría"
    COLORES = {
        "Conciliado": "D9EAD3",
        "Conciliado por fecha": "D9EAD3",
        "Conciliado por referencia": "CFE2F3",
        "Conciliado por agrupacion": "FFF2CC",
        "Hallazgo - Falta en Sistema": "F4CCCC",
        "Hallazgo - Falta en Banco": "FCE5CD",
        "Hallazgo - Monto coincide pero fecha/referencia no": "F9CB9C",
        "Pendiente": "E7E6E6",
    }
    RELLENO_TITULO = PatternFill(start_color="163A5F", end_color="163A5F", fill_type="solid")
    RELLENO_SUBTITULO = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    RELLENO_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    BORDE = Border(
        left=Side(style="thin", color="B7C9D6"),
        right=Side(style="thin", color="B7C9D6"),
        top=Side(style="thin", color="B7C9D6"),
        bottom=Side(style="thin", color="B7C9D6"),
    )

    def exportar(self, df: pd.DataFrame, nombre_archivo: str = "conciliacion.xlsx", metadata: Optional[Dict[str, str]] = None) -> BytesIO:
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"

        metadata = metadata or {}
        df = df.copy()
        if self.ESTADO_COL not in df.columns and "Estado de Auditoria" in df.columns:
            df[self.ESTADO_COL] = df["Estado de Auditoria"]

        self._crear_resumen(ws_resumen, df, metadata, nombre_archivo)
        self._crear_detalle(wb, df, metadata)
        self._crear_hallazgos(wb, df)
        self._agregar_leyenda(wb)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def exportar_resumen(self, df_banco: pd.DataFrame, df_gestion: pd.DataFrame, metadata: Optional[Dict[str, str]] = None) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"

        metadata = metadata or {}
        total_banco = df_banco["Monto"].sum() if "Monto" in df_banco.columns else 0
        total_gestion = df_gestion["Monto"].sum() if "Monto" in df_gestion.columns else 0

        ws["A1"] = "Resumen de conciliacion bancaria"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = self.RELLENO_TITULO
        ws.merge_cells("A1:B1")

        fila = 3
        for clave, valor in self._metadata_items(metadata):
            ws.cell(row=fila, column=1, value=clave)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1

        fila += 1
        indicadores = [
            ("Movimientos banco", len(df_banco)),
            ("Movimientos sistema", len(df_gestion)),
            ("Total banco", total_banco),
            ("Total sistema", total_gestion),
            ("Diferencia", total_banco - total_gestion),
        ]
        for clave, valor in indicadores:
            ws.cell(row=fila, column=1, value=clave)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1

        self._formatear_tabla(ws, 3, fila - 1, 2)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _crear_resumen(self, ws, df: pd.DataFrame, metadata: Dict[str, str], nombre_archivo: str) -> None:
        conciliados = int(df[self.ESTADO_COL].astype(str).str.contains("Conciliado", na=False).sum())
        falta_sistema = int(df[self.ESTADO_COL].astype(str).str.contains("Falta en Sistema", na=False).sum())
        falta_banco = int(df[self.ESTADO_COL].astype(str).str.contains("Falta en Banco", na=False).sum())
        revisar = int(df[self.ESTADO_COL].astype(str).str.contains("fecha/referencia", na=False).sum())
        agrupados = int(df[self.ESTADO_COL].astype(str).str.contains("agrupacion", case=False, na=False).sum())

        ws["A1"] = "CONCILIACION BANCARIA"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = self.RELLENO_TITULO
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:F1")

        ws["A2"] = "Reporte ejecutivo para presentacion a la junta"
        ws["A2"].font = Font(size=11, italic=True, color="163A5F")
        ws.merge_cells("A2:F2")

        fila = 4
        for clave, valor in self._metadata_items(metadata):
            ws.cell(row=fila, column=1, value=clave)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1

        fila += 1
        ws.cell(row=fila, column=1, value="Indicador")
        ws.cell(row=fila, column=2, value="Valor")
        for cell in ws[fila]:
            cell.fill = self.RELLENO_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDE

        resumen = [
            ("Archivo generado", nombre_archivo),
            ("Total registros evaluados", len(df)),
            ("Movimientos conciliados", conciliados),
            ("Conciliados por agrupacion", agrupados),
            ("Hallazgos: falta en sistema", falta_sistema),
            ("Hallazgos: falta en banco", falta_banco),
            ("Partidas por revisar", revisar),
        ]
        inicio = fila + 1
        for clave, valor in resumen:
            fila += 1
            ws.cell(row=fila, column=1, value=clave)
            ws.cell(row=fila, column=2, value=valor)

        self._formatear_tabla(ws, inicio, fila, 2)

        fila += 2
        ws.cell(row=fila, column=1, value="Conclusion ejecutiva")
        ws.cell(row=fila, column=1).fill = self.RELLENO_SUBTITULO
        ws.cell(row=fila, column=1).font = Font(bold=True, color="163A5F")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)

        fila += 1
        mensaje = (
            f"Se conciliaron {conciliados} registros. "
            f"Quedan {falta_sistema + falta_banco + revisar} partidas con observacion "
            "que deben revisarse antes de su presentacion final."
        )
        ws.cell(row=fila, column=1, value=mensaje)
        ws.cell(row=fila, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila + 2, end_column=6)

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22 if col < 3 else 18

    def _crear_detalle(self, wb: Workbook, df: pd.DataFrame, metadata: Dict[str, str]) -> None:
        ws = wb.create_sheet("Detalle")
        headers = [col for col in df.columns if col != "Estado de Auditoria"]

        ws["A1"] = "Detalle de conciliacion"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = self.RELLENO_TITULO
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        subtitulo = " | ".join(f"{k}: {v}" for k, v in self._metadata_items(metadata)[:4])
        ws["A2"] = subtitulo
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = self.RELLENO_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDE

        for row_idx, (_, row) in enumerate(df[headers].iterrows(), header_row + 1):
            estado = str(row.get(self.ESTADO_COL, ""))
            fill = self.COLORES.get(estado)
            for col_idx, header in enumerate(headers, 1):
                valor = row[header]
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = self.BORDE
                if fill:
                    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
                if header in {"Ingreso", "Egreso", "Monto", "Saldo"} and valor not in ("", None):
                    cell.number_format = '#,##0.00'
                if header == "Fecha" and valor not in ("", None):
                    cell.number_format = "DD/MM/YYYY"

        self._autoajustar_columnas(ws)
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = ws.dimensions

    def _crear_hallazgos(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Hallazgos")
        hallazgos = df[~df[self.ESTADO_COL].astype(str).str.contains("Conciliado", na=False)].copy()

        ws["A1"] = "Partidas observadas"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = self.RELLENO_TITULO
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(hallazgos.columns)))

        if hallazgos.empty:
            ws["A3"] = "No hay hallazgos pendientes."
            return

        for col_idx, header in enumerate(hallazgos.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = self.RELLENO_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDE

        for row_idx, (_, row) in enumerate(hallazgos.iterrows(), 4):
            estado = str(row.get(self.ESTADO_COL, ""))
            fill = self.COLORES.get(estado)
            for col_idx, header in enumerate(hallazgos.columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
                cell.border = self.BORDE
                if fill:
                    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
                if header in {"Ingreso", "Egreso", "Monto", "Saldo"} and row[header] not in ("", None):
                    cell.number_format = '#,##0.00'
                if header == "Fecha" and row[header] not in ("", None):
                    cell.number_format = "DD/MM/YYYY"

        self._autoajustar_columnas(ws)
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = ws.dimensions

    def _agregar_leyenda(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Leyenda")
        ws["A1"] = "Leyenda de estados"
        ws["A1"].font = Font(size=15, bold=True, color="FFFFFF")
        ws["A1"].fill = self.RELLENO_TITULO
        ws.merge_cells("A1:C1")

        ws.append(["Estado", "Color", "Descripcion"])
        for cell in ws[2]:
            cell.fill = self.RELLENO_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = self.BORDE

        fila = 3
        for estado, color in self.COLORES.items():
            ws.cell(row=fila, column=1, value=estado)
            ws.cell(row=fila, column=2, value="")
            ws.cell(row=fila, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=fila, column=3, value=self._descripcion_estado(estado))
            for col in range(1, 4):
                ws.cell(row=fila, column=col).border = self.BORDE
            fila += 1

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 55

    def _metadata_items(self, metadata: Dict[str, str]) -> list[tuple[str, str]]:
        orden = [
            ("Administración", metadata.get("Administración", "")),
            ("Edificio", metadata.get("Edificio", "")),
            ("Dirección", metadata.get("Dirección", "")),
            ("Banco", metadata.get("Banco", "")),
            ("Número de cuenta", metadata.get("Número de cuenta", "")),
            ("Moneda", metadata.get("Moneda", "")),
            ("Periodo", metadata.get("Fecha", "")),
            ("Actualizado el", metadata.get("Actualizado el", "")),
        ]
        return [(k, v) for k, v in orden if v]

    def _descripcion_estado(self, estado: str) -> str:
        descripciones = {
            "Conciliado": "Movimiento encontrado con misma fecha y monto.",
            "Conciliado por fecha": "Movimiento encontrado con mismo monto y fecha cercana.",
            "Conciliado por referencia": "Movimiento conciliado por monto e identificadores de descripcion.",
            "Conciliado por agrupacion": "Varios movimientos del sistema suman el monto bancario.",
            "Hallazgo - Falta en Sistema": "Existe en banco y no fue localizado en el sistema.",
            "Hallazgo - Falta en Banco": "Existe en sistema y no fue localizado en el banco.",
            "Hallazgo - Monto coincide pero fecha/referencia no": "Tiene monto similar, pero requiere revision manual.",
            "Pendiente": "Registro aun no procesado.",
        }
        return descripciones.get(estado, "Estado no documentado.")

    def _formatear_tabla(self, ws, fila_inicio: int, fila_fin: int, col_fin: int) -> None:
        for row in range(fila_inicio, fila_fin + 1):
            for col in range(1, col_fin + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDE
                if col == 1:
                    cell.font = Font(bold=True, color="163A5F")

    def _autoajustar_columnas(self, ws) -> None:
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 45)
